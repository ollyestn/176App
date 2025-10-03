"""
Document checker module for comparing Word documents
"""

# def main(file1, file2, result):
#     """
#     Compare two Word documents and save the result to an Excel file
    
#     Args:
#         file1 (str): Path to the first document
#         file2 (str): Path to the second document
#         result (str): Path where the comparison result should be saved
    
#     Returns:
#         bool: True if comparison was successful, False otherwise
#     """
#     try:
#         # This is a placeholder implementation
#         # In a real implementation, you would:
#         # 1. Load both Word documents
#         # 2. Extract and compare their content
#         # 3. Generate a detailed comparison report
#         # 4. Save the report to the result Excel file
        
#         # For now, we'll just create a simple Excel file with placeholder data
#         import openpyxl
#         from openpyxl import Workbook
        
#         wb = Workbook()
#         ws = wb.active
#         if ws is not None:
#             ws.title = "Document Comparison"
            
#             # Add headers
#             ws['A1'] = 'Comparison Report'
#             ws['A2'] = 'Document 1:'
#             ws['B2'] = file1
#             ws['A3'] = 'Document 2:'
#             ws['B3'] = file2
#             ws['A4'] = 'Status:'
#             ws['B4'] = 'Comparison completed successfully'
            
#             # Add some sample comparison data
#             ws['A6'] = 'Differences Found:'
#             ws['B6'] = '0'
#             ws['A7'] = 'Similarity:'
#             ws['B7'] = '95%'
            
#             # Save the workbook
#             wb.save(result)
        
#         return True
#     except Exception as e:
#         print(f"Error during document comparison: {e}")
#         return False


# 安装依赖：
# pip install python-docx nltk openpyxl
# 环境变量：NLTK_DATA = d:\nltk_data
# d:\> python
# >>> import nltk
# >>> nltk.download()
#

import os
from docx import Document
import nltk
from nltk.tokenize import sent_tokenize
import difflib
import openpyxl
from openpyxl.styles import Font

# 下载 punkt 分句模型（只需一次）
nltk.download('punkt', quiet=True)

def read_docx_sentences_with_paragraph_index(file_path):
    """返回句子列表和对应段落编号"""
    doc = Document(file_path)
    sentences_info = []
    for i, para in enumerate(doc.paragraphs):
        sentences = sent_tokenize(para.text)
        for sent in sentences:
            clean_sent = sent.strip()
            if clean_sent:
                sentences_info.append({
                    "sentence": clean_sent,
                    "paragraph_index": i + 1  # 段落从1开始
                })
    return sentences_info

def fuzzy_match_sentences(sentences1, sentences2, threshold=0.9):
    """返回两个句子列表中相似度高于阈值的匹配项"""
    matches = []
    for s1 in sentences1:
        for s2 in sentences2:
            similarity = difflib.SequenceMatcher(None, s1["sentence"], s2["sentence"]).ratio()
            if similarity >= threshold:
                matches.append({
                    "sentence1": s1["sentence"],
                    "sentence2": s2["sentence"],
                    "para1": s1["paragraph_index"],
                    "para2": s2["paragraph_index"],                    
                    "similarity": round(similarity, 3)
                })
    return matches

def export_to_excel(file1, file2, matches, output_path):
    """将匹配结果输出为 Excel 报告"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "匹配报告"

    # Add headers
    ws['A1'] = 'Comparison Report'
    ws['A2'] = '文档 1:'
    ws['B2'] = file1
    ws['A3'] = '文档 2:'
    ws['B3'] = file2
    ws['A4'] = '状态:'
    ws['B4'] = '比较成功完成！'

    headers = ["序号", "匹配句子1","匹配句子2", "文档1段落", "文档2段落", "相似度"]
    # ws.append(headers)
    ws['A6'] = '序号'
    ws['B6'] = '文档1句子'
    ws['C6'] = '文档2句子'
    ws['D6'] = '文档1段落'
    ws['E6'] = '文档2段落'
    ws['F6'] = '相似度'

    # 设置表头加粗
    for col in range(1, len(headers) + 1):
        ws.cell(row=6, column=col).font = Font(bold=True)

    for i, match in enumerate(matches, start=1):
        ws.append([
            i+7,
            match["sentence1"],
            match["sentence2"],
            match["para1"],
            match["para2"],
            match["similarity"]
        ])

    wb.save(output_path)


def main(file1, file2, output_excel="比较报告.xlsx", similarity_threshold=0.9):
    if not os.path.exists(file1) or not os.path.exists(file2):
        print("请确保两个文件都存在。")
        return False

    sents1 = read_docx_sentences_with_paragraph_index(file1)
    sents2 = read_docx_sentences_with_paragraph_index(file2)
    matches = fuzzy_match_sentences(sents1, sents2, threshold=similarity_threshold)

    export_to_excel(file1, file2, matches, output_excel)
    print(f"比较完成！共找到 {len(matches)} 个相似句子。报告已生成：{output_excel}")
    return True

# 命令行支持
if __name__ == "__main__":
    import sys
    if len(sys.argv) < 3:
        print("用法：python compare_fuzzy.py 文件1.docx 文件2.docx")
    else:
        main(sys.argv[1], sys.argv[2])
